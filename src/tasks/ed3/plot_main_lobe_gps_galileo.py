from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import csv
import math
import re
from zipfile import ZipFile

import matplotlib.pyplot as plt
from matplotlib import patches
from matplotlib.gridspec import GridSpecFromSubplotSpec
from matplotlib.lines import Line2D
from matplotlib.offsetbox import AnnotationBbox, OffsetImage
import numpy as np
import openpyxl
import pandas as pd
from pypdf import PdfReader

from plot_main_lobe_grap_revised import load_grap_band


ROOT = Path(__file__).resolve().parent
GPS_DIR = ROOT / "GPS_antenna_patterns"
GRAP_DIR = ROOT / "GRAP_metadata"
WINDOWS_CSV = ROOT / "lugre_full_geometry_noise_windows_refined.csv"
IGS_METADATA = GPS_DIR / "igs_satellite_metadata.snx"
SATELLITE_IMAGE = ROOT / "gps_iii_satellite_cutout.png"
EARTH_IMAGE = ROOT / "earth_user_reference_globe.png"
OUTPUT_STEM = ROOT / "lugre_main_lobe_ab_figure_nature_gps_galileo"
CURRENT_STEM = ROOT / "lugre_main_lobe_ab_figure_nature"

SIGNALS = ["GPS L1", "GPS L5", "Galileo E1", "Galileo E5a"]
SIGNAL_LABELS = {
    "GPS L1": "GPS L1 C/A",
    "GPS L5": "GPS L5Q",
    "Galileo E1": "Galileo E1C",
    "Galileo E5a": "Galileo E5a-Q",
}
SIGNAL_FREQUENCIES = {
    "GPS L1": "1575.42 MHz",
    "GPS L5": "1176.45 MHz",
    "Galileo E1": "1575.42 MHz",
    "Galileo E5a": "1176.45 MHz",
}
COLORS = {
    "GPS L1": "#2D6C9F",
    "GPS L5": "#69A6D3",
    "Galileo E1": "#E2A532",
    "Galileo E5a": "#D96B43",
    "ink": "#1F2833",
    "muted": "#66717D",
    "grid": "#DCE3E8",
    "earth_limb": "#7B858F",
    "gps_limb": "#526F8A",
    "galileo_limb": "#A47C25",
    "observation": "#008A78",
    "main": "#7A5AA6",
    "side": "#B54F56",
    "schematic_main": "#B9A4D2",
    "schematic_side": "#D9B4C2",
}

TARGET_EPOCH = datetime(2025, 3, 1, tzinfo=timezone.utc)
THETA_GRID = np.arange(0.0, 76.0, 1.0)
EARTH_LIMB_DEG = {"GPS": 13.8, "Galileo": 12.4}
EARTH_LIMB_STYLE = {"GPS": (0, (3.0, 2.4)), "Galileo": (0, (1.2, 1.8))}
EARTH_LIMB_COLOR = {
    "GPS": COLORS["gps_limb"],
    "Galileo": COLORS["galileo_limb"],
}


@dataclass
class PatternSummary:
    theta: np.ndarray
    median: np.ndarray
    p05: np.ndarray
    p95: np.ndarray
    n_sv: np.ndarray
    svns: list[int]
    source_note: str
    model_lower: np.ndarray | None = None
    model_upper: np.ndarray | None = None


def configure_style() -> None:
    plt.rcParams.update(
        {
            "font.family": "Arial",
            "font.size": 7.2,
            "axes.titlesize": 7.8,
            "axes.labelsize": 7.0,
            "xtick.labelsize": 6.4,
            "ytick.labelsize": 6.4,
            "legend.fontsize": 6.0,
            "axes.linewidth": 0.7,
            "axes.edgecolor": COLORS["ink"],
            "axes.labelcolor": COLORS["ink"],
            "xtick.color": COLORS["ink"],
            "ytick.color": COLORS["ink"],
            "xtick.direction": "out",
            "ytick.direction": "out",
            "xtick.major.size": 2.5,
            "ytick.major.size": 2.5,
            "savefig.facecolor": "white",
            "figure.facecolor": "white",
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
        }
    )


def style_axis(ax: plt.Axes) -> None:
    ax.grid(True, color=COLORS["grid"], linewidth=0.45, alpha=0.82)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_linewidth(0.65)
        spine.set_color(COLORS["ink"])


def load_windows() -> tuple[pd.DataFrame, pd.DataFrame]:
    usecols = [
        "phase",
        "op_id",
        "svid",
        "sat",
        "signal_name",
        "window_start_gps_seconds",
        "window_end_gps_seconds",
        "sigma_ratio_to_main_anchor",
        "tx_off_boresight_deg",
        "evidence_class",
    ]
    raw = pd.read_csv(WINDOWS_CSV, usecols=usecols)
    raw = raw[raw["signal_name"].isin(SIGNALS)].copy()
    numeric = [
        "window_start_gps_seconds",
        "window_end_gps_seconds",
        "sigma_ratio_to_main_anchor",
        "tx_off_boresight_deg",
    ]
    for column in numeric:
        raw[column] = pd.to_numeric(raw[column], errors="coerce")
    raw = raw.dropna(subset=numeric)
    raw = raw[
        (raw["sigma_ratio_to_main_anchor"] > 0)
        & raw["tx_off_boresight_deg"].between(0, 90)
    ].copy()

    # The source windows advance every 60 s but span 120 s. Keep the first
    # non-overlapping sequence within each operation/SV/signal combination.
    keep: list[int] = []
    keys = ["phase", "op_id", "svid", "signal_name"]
    for _, group in raw.sort_values("window_start_gps_seconds").groupby(keys, sort=False):
        last_end = -np.inf
        for idx, row in group.iterrows():
            if row["window_start_gps_seconds"] > last_end:
                keep.append(idx)
                last_end = row["window_end_gps_seconds"]
    independent = raw.loc[keep].sort_values("window_start_gps_seconds").copy()
    return raw, independent


def parse_sinex_epoch(text: str, is_end: bool = False) -> datetime:
    year_text, day_text, second_text = text.split(":")
    year = int(year_text)
    if year == 0:
        return datetime.max.replace(tzinfo=timezone.utc) if is_end else datetime.min.replace(
            tzinfo=timezone.utc
        )
    return datetime(year, 1, 1, tzinfo=timezone.utc) + timedelta(
        days=int(day_text) - 1, seconds=int(second_text)
    )


def load_prn_svn_map(epoch: datetime = TARGET_EPOCH) -> dict[str, int]:
    mapping: dict[str, int] = {}
    in_block = False
    for line in IGS_METADATA.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.startswith("+SATELLITE/PRN"):
            in_block = True
            continue
        if line.startswith("-SATELLITE/PRN"):
            break
        if not in_block or not line.startswith(" G"):
            continue
        fields = line.split()
        if len(fields) < 4:
            continue
        svn_text, start_text, end_text, prn = fields[:4]
        start = parse_sinex_epoch(start_text)
        end = parse_sinex_epoch(end_text, is_end=True)
        if start <= epoch < end:
            mapping[prn] = int(svn_text[1:])
    return mapping


def _split_signed_cut(theta: np.ndarray, gain: np.ndarray) -> list[tuple[np.ndarray, np.ndarray]]:
    profiles: list[tuple[np.ndarray, np.ndarray]] = []
    for mask in (theta >= 0, theta <= 0):
        cut_theta = np.abs(theta[mask])
        cut_gain = gain[mask]
        order = np.argsort(cut_theta)
        cut_theta = cut_theta[order]
        cut_gain = cut_gain[order]
        unique_theta, unique_indices = np.unique(cut_theta, return_index=True)
        cut_gain = cut_gain[unique_indices]
        valid = np.isfinite(unique_theta) & np.isfinite(cut_gain)
        if valid.sum() >= 8:
            profiles.append((unique_theta[valid], cut_gain[valid]))
    return profiles


def load_iir_patterns(
    target_svns: set[int], signal: str
) -> dict[int, list[tuple[np.ndarray, np.ndarray]]]:
    if signal != "GPS L1":
        return {}
    path = GPS_DIR / "GPS_IIR_IIRM_SV_specific_patterns_data.pptx"
    output: dict[int, list[tuple[np.ndarray, np.ndarray]]] = {}
    with ZipFile(path) as archive:
        workbook_names = [
            name
            for name in archive.namelist()
            if name.startswith("ppt/embeddings/") and name.lower().endswith(".xlsx")
        ]
        workbook_names.sort(
            key=lambda name: int(re.search(r"(\d+)\.xlsx$", name).group(1))
        )
        for name in workbook_names:
            workbook = openpyxl.load_workbook(
                BytesIO(archive.read(name)), read_only=True, data_only=True
            )
            data_sheets = [
                sheet_name for sheet_name in workbook.sheetnames if "data" in sheet_name.lower()
            ]
            if not data_sheets:
                workbook.close()
                continue
            match = re.search(r"SVN(\d+)\s+L(\d)", data_sheets[0], flags=re.IGNORECASE)
            if not match:
                workbook.close()
                continue
            svn = int(match.group(1))
            band = match.group(2)
            if svn not in target_svns or band != "1":
                workbook.close()
                continue
            sheet = workbook[data_sheets[0]]
            rows = list(sheet.iter_rows(values_only=True))
            phi = np.asarray(rows[1][1:], dtype=float)
            theta = np.asarray([row[0] for row in rows[3:] if row[0] is not None], dtype=float)
            matrix = np.asarray(
                [
                    [np.nan if value is None else float(value) for value in row[1 : 1 + len(phi)]]
                    for row in rows[3 : 3 + len(theta)]
                ],
                dtype=float,
            )
            profiles: list[tuple[np.ndarray, np.ndarray]] = []
            for column in range(matrix.shape[1]):
                profiles.extend(_split_signed_cut(theta, matrix[:, column]))
            output[svn] = profiles
            workbook.close()
    return output


def _find_numeric_blocks(
    rows: list[tuple], target_frequency: float, mode: str
) -> list[tuple[float, np.ndarray, np.ndarray]]:
    blocks: list[tuple[float, np.ndarray, np.ndarray]] = []
    for idx, row in enumerate(rows):
        if (
            len(row) < 2
            or str(row[0]).upper() != "FREQ"
            or not isinstance(row[1], (int, float))
            or not math.isclose(float(row[1]), target_frequency, abs_tol=5e-5)
        ):
            continue
        if idx + 2 >= len(rows):
            continue
        descriptor = rows[idx + 1]
        descriptor_text = " ".join(str(value).upper() for value in descriptor[:4])
        if mode == "great" and "GREAT" not in descriptor_text:
            continue
        if mode == "conical" and "CONICAL" not in descriptor_text:
            continue
        cut_value = next(
            (
                float(value)
                for value in descriptor[2:5]
                if isinstance(value, (int, float))
            ),
            np.nan,
        )
        angles: list[float] = []
        gains: list[float] = []
        for data_row in rows[idx + 3 :]:
            if (
                len(data_row) < 2
                or not isinstance(data_row[0], (int, float))
                or not isinstance(data_row[1], (int, float))
            ):
                break
            angles.append(float(data_row[0]))
            gains.append(float(data_row[1]))
        if len(angles) >= 8:
            blocks.append((cut_value, np.asarray(angles), np.asarray(gains)))
    return blocks


def load_iif_patterns(
    target_svns: set[int], signal: str
) -> dict[int, list[tuple[np.ndarray, np.ndarray]]]:
    target_frequency = 1.57542 if signal == "GPS L1" else 1.17645
    output: dict[int, list[tuple[np.ndarray, np.ndarray]]] = {}
    for svn in sorted(target_svns):
        path = GPS_DIR / f"GPS_IIF_SVN{svn}_L1_L2_L5.xlsx"
        if not path.exists():
            continue
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        great_sheet = next(
            (
                name
                for name in workbook.sheetnames
                if name.strip().upper() in {"LBS GC", "GCOUT"}
            ),
            None,
        )
        profiles: list[tuple[np.ndarray, np.ndarray]] = []
        if great_sheet is not None:
            rows = list(workbook[great_sheet].iter_rows(values_only=True))
            for _, theta, gain in _find_numeric_blocks(rows, target_frequency, "great"):
                profiles.extend(_split_signed_cut(theta, gain))
        else:
            rows = list(workbook["LBSCC"].iter_rows(values_only=True))
            blocks = _find_numeric_blocks(rows, target_frequency, "conical")
            theta_cuts = np.asarray([block[0] for block in blocks], dtype=float)
            order = np.argsort(theta_cuts)
            theta_cuts = theta_cuts[order]
            blocks = [blocks[index] for index in order]
            target_phi = np.arange(0.0, 360.0, 10.0)
            matrix = np.full((len(target_phi), len(blocks)), np.nan, dtype=float)
            for column, (_, phi, gain) in enumerate(blocks):
                phi = np.mod(phi, 360.0)
                sort_index = np.argsort(phi)
                phi = phi[sort_index]
                gain = gain[sort_index]
                unique_phi, unique_index = np.unique(phi, return_index=True)
                gain = gain[unique_index]
                phi_extended = np.r_[unique_phi, unique_phi[0] + 360.0]
                gain_extended = np.r_[gain, gain[0]]
                matrix[:, column] = np.interp(target_phi, phi_extended, gain_extended)
            for row in matrix:
                valid = np.isfinite(theta_cuts) & np.isfinite(row)
                if valid.sum() >= 5:
                    profiles.append((theta_cuts[valid], row[valid]))
        workbook.close()
        if profiles:
            output[svn] = profiles
    return output


def _parse_gps_iii_pdf(pdf_bytes: bytes) -> list[tuple[np.ndarray, np.ndarray]]:
    text = PdfReader(BytesIO(pdf_bytes)).pages[0].extract_text()
    theta_rows: list[float] = []
    matrix_rows: list[list[float]] = []
    for line in text.splitlines():
        fields = line.split()
        if len(fields) != 37:
            continue
        try:
            values = [float(value) for value in fields]
        except ValueError:
            continue
        if -90 <= values[0] <= 90:
            theta_rows.append(values[0])
            matrix_rows.append(values[1:])
    theta = np.asarray(theta_rows, dtype=float)
    matrix = np.asarray(matrix_rows, dtype=float)
    profiles: list[tuple[np.ndarray, np.ndarray]] = []
    if matrix.shape[1] == 36:
        for column in range(36):
            profiles.extend(_split_signed_cut(theta, matrix[:, column]))
    return profiles


def load_gps_iii_patterns(
    target_svns: set[int], signal: str
) -> dict[int, list[tuple[np.ndarray, np.ndarray]]]:
    band_token = "_L1_" if signal == "GPS L1" else "_L5_"
    output: dict[int, list[tuple[np.ndarray, np.ndarray]]] = {}
    for svn in sorted(target_svns):
        zip_path = GPS_DIR / f"GPS_III_SVN{svn}_Directivity.zip"
        if not zip_path.exists():
            continue
        with ZipFile(zip_path) as archive:
            pdf_name = next(
                (
                    name
                    for name in archive.namelist()
                    if band_token in name.upper() and name.lower().endswith(".pdf")
                ),
                None,
            )
            if pdf_name is None:
                continue
            profiles = _parse_gps_iii_pdf(archive.read(pdf_name))
        if profiles:
            output[svn] = profiles
    return output


def _safe_nanquantile(matrix: np.ndarray, quantile: float) -> np.ndarray:
    result = np.full(matrix.shape[1], np.nan, dtype=float)
    for column in range(matrix.shape[1]):
        values = matrix[:, column]
        values = values[np.isfinite(values)]
        if values.size:
            result[column] = float(np.quantile(values, quantile))
    return result


def summarize_satellite_patterns(
    patterns: dict[int, list[tuple[np.ndarray, np.ndarray]]],
    source_note: str,
) -> PatternSummary:
    per_sv_quantiles: list[np.ndarray] = []
    n_sv = np.zeros_like(THETA_GRID, dtype=int)
    for svn in sorted(patterns):
        interpolated: list[np.ndarray] = []
        for theta, gain in patterns[svn]:
            valid = np.isfinite(theta) & np.isfinite(gain)
            if valid.sum() < 5:
                continue
            theta = theta[valid]
            gain = gain[valid]
            order = np.argsort(theta)
            theta = theta[order]
            gain = gain[order]
            zero = float(np.interp(0.0, theta, gain))
            profile = np.interp(THETA_GRID, theta, gain - zero, left=np.nan, right=np.nan)
            profile[THETA_GRID > np.nanmax(theta)] = np.nan
            interpolated.append(profile)
        if not interpolated:
            continue
        matrix = np.vstack(interpolated)
        n_sv += np.any(np.isfinite(matrix), axis=0).astype(int)
        quantile_profiles = np.vstack(
            [
                _safe_nanquantile(matrix, quantile)
                for quantile in np.linspace(0.05, 0.95, 19)
            ]
        )
        per_sv_quantiles.append(quantile_profiles)
    if not per_sv_quantiles:
        raise RuntimeError(f"No satellite patterns were loaded for {source_note}")
    equal_weight_samples = np.vstack(per_sv_quantiles)
    return PatternSummary(
        theta=THETA_GRID,
        median=_safe_nanquantile(equal_weight_samples, 0.50),
        p05=_safe_nanquantile(equal_weight_samples, 0.05),
        p95=_safe_nanquantile(equal_weight_samples, 0.95),
        n_sv=n_sv,
        svns=sorted(patterns),
        source_note=source_note,
    )


def load_gps_summary(
    signal: str, observed_prns: set[str], prn_svn: dict[str, int]
) -> tuple[PatternSummary, list[int], list[int]]:
    observed_svns = sorted(
        {prn_svn[prn] for prn in observed_prns if prn in prn_svn}
    )
    target_svns = set(observed_svns)
    patterns: dict[int, list[tuple[np.ndarray, np.ndarray]]] = {}
    patterns.update(load_iir_patterns(target_svns & set(range(41, 62)), signal))
    patterns.update(load_iif_patterns(target_svns & set(range(62, 74)), signal))
    patterns.update(load_gps_iii_patterns(target_svns & set(range(74, 79)), signal))
    missing = sorted(target_svns - set(patterns))
    summary = summarize_satellite_patterns(
        patterns,
        "NAVCEN measured SV patterns; equal weight per patterned SVN",
    )
    return summary, observed_svns, missing


def load_galileo_summary(signal: str) -> PatternSummary:
    filename = "GRAP_File_E1_.xlsx" if signal == "Galileo E1" else "GRAP_File_E5a_.xlsx"
    grap = load_grap_band(GRAP_DIR / filename)
    theta = grap["theta"]
    keep = theta <= THETA_GRID[-1]
    theta = theta[keep]
    eirp = grap["eirp"][:, keep]
    upper = grap["upper"][:, keep]
    lower = grap["lower"][:, keep]
    eirp_norm = eirp - eirp[:, [0]]
    # Confidence limits remain relative to the GRAP median boresight value.
    upper_norm = upper - eirp[:, [0]]
    lower_norm = lower - eirp[:, [0]]
    return PatternSummary(
        theta=theta,
        median=np.nanmedian(eirp_norm, axis=0),
        p05=np.nanquantile(eirp_norm, 0.05, axis=0),
        p95=np.nanquantile(eirp_norm, 0.95, axis=0),
        n_sv=np.full_like(theta, np.nan, dtype=float),
        svns=[],
        source_note="JRC GRAP Galileo FOC population model",
        model_lower=np.nanmedian(lower_norm, axis=0),
        model_upper=np.nanmedian(upper_norm, axis=0),
    )


def binned_distribution(
    x: np.ndarray, y: np.ndarray, bins: np.ndarray
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    centers = 0.5 * (bins[:-1] + bins[1:])
    p05 = np.full_like(centers, np.nan, dtype=float)
    median = np.full_like(centers, np.nan, dtype=float)
    p95 = np.full_like(centers, np.nan, dtype=float)
    for index in range(len(centers)):
        values = y[(x >= bins[index]) & (x < bins[index + 1])]
        values = values[np.isfinite(values)]
        if values.size >= 8:
            p05[index], median[index], p95[index] = np.quantile(
                values, [0.05, 0.50, 0.95]
            )
    return centers, p05, median, p95


def draw_panel_a(
    fig: plt.Figure, subplot_spec, windows: pd.DataFrame
) -> tuple[list[plt.Axes], list[Line2D | patches.Patch]]:
    grid = GridSpecFromSubplotSpec(
        2, 2, subplot_spec=subplot_spec, hspace=0.24, wspace=0.19
    )
    axes: list[plt.Axes] = []
    bins = np.arange(10.0, 76.0 + 2.0, 2.0)
    for index, signal in enumerate(SIGNALS):
        ax = fig.add_subplot(grid[index // 2, index % 2])
        axes.append(ax)
        subset = windows[windows["signal_name"] == signal]
        x = subset["tx_off_boresight_deg"].to_numpy(dtype=float)
        y = subset["sigma_ratio_to_main_anchor"].to_numpy(dtype=float)
        color = COLORS[signal]
        ax.scatter(
            x,
            y,
            s=5.0,
            color=color,
            alpha=0.19,
            linewidths=0,
            rasterized=True,
            zorder=2,
        )
        centers, p05, median, p95 = binned_distribution(x, y, bins)
        valid = np.isfinite(median)
        ax.fill_between(
            centers,
            p05,
            p95,
            where=valid,
            color=color,
            alpha=0.17,
            linewidth=0,
            zorder=3,
        )
        ax.plot(
            centers[valid],
            median[valid],
            color=color,
            linewidth=1.45,
            zorder=4,
        )

        for evidence, marker, evidence_color in [
            ("main-lobe anchor", "o", COLORS["main"]),
            ("side-lobe anchor", "x", COLORS["side"]),
        ]:
            selected = subset[subset["evidence_class"] == evidence]
            if selected.empty:
                continue
            scatter_kwargs: dict[str, object] = {
                "s": 15,
                "marker": marker,
                "linewidths": 0.75,
                "zorder": 5,
            }
            if marker == "o":
                scatter_kwargs.update(
                    {"facecolors": "none", "edgecolors": evidence_color}
                )
            else:
                scatter_kwargs["color"] = evidence_color
            ax.scatter(
                selected["tx_off_boresight_deg"],
                selected["sigma_ratio_to_main_anchor"],
                **scatter_kwargs,
            )

        ax.axhline(1.0, color=COLORS["ink"], linewidth=0.65, linestyle=(0, (4, 3)))
        ax.axvline(
            13.8,
            color=COLORS["earth_limb"],
            linewidth=0.75,
            linestyle=(0, (2, 2)),
            zorder=1,
        )
        ax.set_yscale("log")
        ax.set_xlim(10, 75)
        ax.set_ylim(0.45, 16)
        ax.set_yticks([0.5, 1, 2, 5, 10])
        ax.set_yticklabels(["0.5", "1", "2", "5", "10"])
        ax.set_title(
            f"{SIGNAL_LABELS[signal]}  |  {SIGNAL_FREQUENCIES[signal]}",
            loc="left",
            color=color,
            fontweight="bold",
            pad=4,
        )
        ax.text(
            0.98,
            0.94,
            f"n={len(subset):,}; {subset['sat'].nunique()} PRNs",
            transform=ax.transAxes,
            ha="right",
            va="top",
            color=COLORS["muted"],
            fontsize=5.8,
        )
        if index // 2 == 1:
            ax.set_xlabel(r"Off-boresight angle, $\theta$ (deg)")
        else:
            ax.tick_params(labelbottom=False)
        if index % 2 == 0:
            ax.set_ylabel(r"$\sigma_{\rho}/\widehat{\sigma}_{\rho,\,main}$")
        else:
            ax.tick_params(labelleft=False)
        style_axis(ax)

    legend = [
        Line2D(
            [0],
            [0],
            marker="o",
            color="none",
            markerfacecolor=COLORS["muted"],
            markeredgewidth=0,
            alpha=0.35,
            markersize=4,
            label="All non-overlapping 120-s windows",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            color="none",
            markerfacecolor="none",
            markeredgecolor=COLORS["main"],
            markeredgewidth=0.8,
            markersize=4,
            label="Paper main-lobe anchor",
        ),
        Line2D(
            [0],
            [0],
            marker="x",
            color=COLORS["side"],
            linewidth=0,
            markeredgewidth=0.8,
            markersize=4,
            label="Paper side-lobe anchor",
        ),
        Line2D([0], [0], color=COLORS["ink"], linewidth=1.3, label="2-deg-bin median"),
        patches.Patch(
            facecolor=COLORS["muted"],
            alpha=0.18,
            edgecolor="none",
            label="Within-bin 5-95%",
        ),
        Line2D(
            [0],
            [0],
            color=COLORS["earth_limb"],
            linestyle=(0, (2, 2)),
            label="Earth limb, 13.8 deg",
        ),
    ]
    return axes, legend


def draw_image(
    ax: plt.Axes,
    image_path: Path,
    center: tuple[float, float],
    zoom: float,
    zorder: int,
) -> None:
    image = plt.imread(image_path)
    ax.add_artist(
        AnnotationBbox(
            OffsetImage(image, zoom=zoom),
            center,
            frameon=False,
            pad=0,
            zorder=zorder,
        )
    )


def ray_endpoint(
    origin: tuple[float, float], length: float, angle_from_down_deg: float
) -> tuple[float, float]:
    angle = math.radians(angle_from_down_deg)
    return origin[0] + length * math.sin(angle), origin[1] - length * math.cos(angle)


def ray_polygon(
    origin: tuple[float, float],
    length: float,
    left_angle: float,
    right_angle: float,
) -> np.ndarray:
    left = ray_endpoint(origin, length, left_angle)
    right = ray_endpoint(origin, length, right_angle)
    return np.asarray([origin, left, right])


def draw_panel_b(fig: plt.Figure, subplot_spec) -> plt.Axes:
    ax = fig.add_subplot(subplot_spec)
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 10)
    ax.axis("off")

    origin = (5.05, 8.55)
    earth_center = (5.05, 2.15)
    earth_radius = (origin[1] - earth_center[1]) * math.sin(math.radians(13.8))

    # Stylized lobe fields only explain the geometry; measured boundaries are in c.
    ax.add_patch(
        patches.Polygon(
            ray_polygon(origin, 8.0, -22.0, 22.0),
            closed=True,
            facecolor=COLORS["schematic_main"],
            edgecolor="none",
            alpha=0.55,
            zorder=1,
        )
    )
    ax.add_patch(
        patches.Polygon(
            ray_polygon(origin, 7.7, -42.0, -25.5),
            closed=True,
            facecolor=COLORS["schematic_side"],
            edgecolor="none",
            alpha=0.35,
            zorder=0,
        )
    )
    ax.add_patch(
        patches.Polygon(
            ray_polygon(origin, 7.7, 25.5, 42.0),
            closed=True,
            facecolor=COLORS["schematic_side"],
            edgecolor="none",
            alpha=0.35,
            zorder=0,
        )
    )

    center_end = ray_endpoint(origin, 7.5, 0)
    limb_end = ray_endpoint(origin, 7.5, 13.8)
    ax.plot(
        [origin[0], center_end[0]],
        [origin[1], center_end[1]],
        color=COLORS["earth_limb"],
        linewidth=0.8,
        linestyle=(0, (4, 3)),
        zorder=4,
    )
    ax.plot(
        [origin[0], limb_end[0]],
        [origin[1], limb_end[1]],
        color=COLORS["earth_limb"],
        linewidth=0.8,
        linestyle=(0, (4, 3)),
        zorder=4,
    )

    arc_radius = 2.55
    arc = patches.Arc(
        origin,
        2 * arc_radius,
        2 * arc_radius,
        angle=0,
        theta1=270,
        theta2=283.8,
        color=COLORS["earth_limb"],
        linewidth=1.0,
        zorder=6,
    )
    ax.add_patch(arc)
    arc_start = ray_endpoint(origin, arc_radius, 0)
    arc_end = ray_endpoint(origin, arc_radius, 13.8)
    ax.annotate(
        "",
        xy=arc_end,
        xytext=arc_start,
        arrowprops=dict(
            arrowstyle="<->",
            color=COLORS["earth_limb"],
            linewidth=0.9,
            shrinkA=0,
            shrinkB=0,
            connectionstyle="arc3,rad=-0.11",
        ),
        zorder=7,
    )

    draw_image(ax, EARTH_IMAGE, earth_center, zoom=0.43, zorder=5)
    draw_image(ax, SATELLITE_IMAGE, (5.05, 9.05), zoom=0.105, zorder=8)

    ax.text(
        6.25,
        9.17,
        "GNSS satellite",
        ha="left",
        va="center",
        fontsize=7.2,
        fontweight="bold",
        color=COLORS["ink"],
    )
    ax.text(
        5.75,
        6.35,
        r"Earth-limb angle  $\alpha=13.8^\circ$",
        ha="left",
        va="center",
        fontsize=6.7,
        color=COLORS["ink"],
    )
    ax.text(
        8.18,
        5.62,
        "Side-lobe\nregion",
        ha="center",
        va="center",
        color="#8B5068",
        fontsize=6.6,
        fontweight="bold",
    )
    ax.text(
        3.35,
        6.80,
        "Central-lobe region\n(pattern dependent)",
        ha="center",
        va="center",
        color="#5B437D",
        fontsize=6.6,
        fontweight="bold",
    )

    lugre = (4.12, 4.96)
    ax.scatter(
        [lugre[0]],
        [lugre[1]],
        s=24,
        color="#B23675",
        edgecolor="white",
        linewidth=0.65,
        zorder=9,
    )
    ax.annotate(
        "LuGRE",
        xy=lugre,
        xytext=(1.95, 5.50),
        ha="left",
        va="center",
        fontsize=7.0,
        fontweight="bold",
        color=COLORS["ink"],
        arrowprops=dict(
            arrowstyle="-",
            color=COLORS["ink"],
            linewidth=0.7,
            shrinkA=2,
            shrinkB=3,
        ),
        zorder=10,
    )
    ax.text(
        5.05,
        0.22,
        "Schematic only: lobe shading is not an angle classifier",
        ha="center",
        va="bottom",
        fontsize=5.7,
        color=COLORS["muted"],
    )
    return ax


def draw_pattern_axis(
    ax: plt.Axes,
    signal: str,
    summary: PatternSummary,
    observations: pd.DataFrame,
) -> None:
    color = COLORS[signal]
    ax.fill_between(
        summary.theta,
        summary.p05,
        summary.p95,
        color=color,
        alpha=0.22,
        linewidth=0,
        zorder=2,
    )
    ax.plot(
        summary.theta,
        summary.median,
        color=color,
        linewidth=1.55,
        zorder=4,
    )

    angles = observations["tx_off_boresight_deg"].to_numpy(dtype=float)
    angle_min = float(np.nanmin(angles))
    angle_max = float(np.nanmax(angles))
    ax.axvspan(
        angle_min,
        angle_max,
        facecolor=COLORS["observation"],
        alpha=0.12,
        linewidth=0,
        zorder=0,
    )
    ax.text(
        0.5 * (angle_min + angle_max),
        -13.0,
        "LuGRE",
        rotation=0,
        ha="center",
        va="center",
        fontsize=6.2,
        fontweight="bold",
        color=COLORS["observation"],
        zorder=5,
    )

    constellation = "GPS" if signal.startswith("GPS") else "Galileo"
    ax.axvline(
        EARTH_LIMB_DEG[constellation],
        color=EARTH_LIMB_COLOR[constellation],
        linewidth=0.85,
        linestyle=EARTH_LIMB_STYLE[constellation],
        zorder=1,
    )

    ax.set_xlim(0, 75)
    ax.set_ylim(-45, 4.4)
    ax.set_yticks([-40, -30, -20, -10, 0])
    ax.set_title(
        SIGNAL_LABELS[signal],
        loc="left",
        color=color,
        fontweight="bold",
        pad=3,
        fontsize=7.0,
    )
    style_axis(ax)


def draw_panel_c(
    fig: plt.Figure,
    subplot_spec,
    summaries: dict[str, PatternSummary],
    windows: pd.DataFrame,
) -> tuple[list[plt.Axes], list[Line2D | patches.Patch]]:
    grid = GridSpecFromSubplotSpec(
        2, 2, subplot_spec=subplot_spec, hspace=0.28, wspace=0.22
    )
    axes: list[plt.Axes] = []
    for index, signal in enumerate(SIGNALS):
        ax = fig.add_subplot(grid[index // 2, index % 2])
        axes.append(ax)
        draw_pattern_axis(
            ax,
            signal,
            summaries[signal],
            windows[windows["signal_name"] == signal],
        )
        if index // 2 == 1:
            ax.set_xlabel(r"Off-boresight angle, $\theta$ (deg)")
        else:
            ax.tick_params(labelbottom=False)
        if index % 2 == 0:
            ax.set_ylabel(r"Relative EIRP, $\Delta$EIRP (dB)")
        else:
            ax.tick_params(labelleft=False)

    legend = [
        Line2D([0], [0], color=COLORS["ink"], linewidth=1.5, label="Pattern median"),
        patches.Patch(
            facecolor=COLORS["muted"],
            alpha=0.22,
            edgecolor="none",
            label="Pattern central 90% (P5-P95)",
        ),
        patches.Patch(
            facecolor=COLORS["observation"],
            alpha=0.12,
            edgecolor="none",
            label="LuGRE observed range (min-max)",
        ),

        Line2D(
            [0],
            [0],
            color=EARTH_LIMB_COLOR["GPS"],
            linestyle=EARTH_LIMB_STYLE["GPS"],
            linewidth=0.85,
            label="GPS limb, 13.8 deg",
        ),
        Line2D(
            [0],
            [0],
            color=EARTH_LIMB_COLOR["Galileo"],
            linestyle=EARTH_LIMB_STYLE["Galileo"],
            linewidth=0.85,
            label="Galileo limb, 12.4 deg",
        ),

    ]
    return axes, legend


def write_coverage(
    raw_windows: pd.DataFrame,
    independent_windows: pd.DataFrame,
    prn_svn: dict[str, int],
    summaries: dict[str, PatternSummary],
    gps_observed_svns: dict[str, list[int]],
    gps_missing_svns: dict[str, list[int]],
) -> Path:
    path = ROOT / "lugre_main_lobe_pattern_coverage.csv"
    rows: list[dict[str, str | int]] = []
    for signal in SIGNALS:
        raw_subset = raw_windows[raw_windows["signal_name"] == signal]
        independent_subset = independent_windows[
            independent_windows["signal_name"] == signal
        ]
        summary = summaries[signal]
        if signal.startswith("GPS"):
            observed_svns = gps_observed_svns[signal]
            patterned = summary.svns
            missing = gps_missing_svns[signal]
            n_at_20 = int(summary.n_sv[np.where(summary.theta == 20)[0][0]])
            n_at_40 = int(summary.n_sv[np.where(summary.theta == 40)[0][0]])
        else:
            observed_svns = []
            patterned = []
            missing = []
            n_at_20 = ""
            n_at_40 = ""
        rows.append(
            {
                "signal": signal,
                "raw_120s_windows": len(raw_subset),
                "nonoverlapping_120s_windows": len(independent_subset),
                "observed_prns": raw_subset["sat"].nunique(),
                "observed_svns": " ".join(f"SVN{value}" for value in observed_svns),
                "patterned_svns": " ".join(f"SVN{value}" for value in patterned),
                "missing_pattern_svns": " ".join(f"SVN{value}" for value in missing),
                "pattern_sv_count_at_20deg": n_at_20,
                "pattern_sv_count_at_40deg": n_at_40,
                "pattern_source": summary.source_note,
            }
        )
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_caption_and_methods(
    raw_windows: pd.DataFrame,
    independent_windows: pd.DataFrame,
    gps_missing_svns: dict[str, list[int]],
) -> tuple[Path, Path]:
    caption_path = ROOT / "lugre_main_lobe_ab_figure_nature_gps_galileo_caption.txt"
    methods_path = ROOT / "lugre_main_lobe_ab_figure_nature_gps_galileo_references_and_methods.txt"

    counts = {
        signal: len(independent_windows[independent_windows["signal_name"] == signal])
        for signal in SIGNALS
    }
    caption = f"""Extended Data Fig. 3 | LuGRE reception geometry and signal-specific transmit-pattern evidence.

a, Normalized pseudorange-noise standard deviation as a function of transmitter off-boresight angle for GPS L1 C/A (n={counts['GPS L1']:,}), GPS L5Q (n={counts['GPS L5']:,}), Galileo E1C (n={counts['Galileo E1']:,}) and Galileo E5a-Q (n={counts['Galileo E5a']:,}). Points are all valid, non-overlapping 120-s windows. Solid curves and shaded bands show the median and 5-95% interval within 2-deg angle bins. No observations were excluded using the 5-95% intervals; these intervals only visualize the observed spread.

b, Definition of the transmitter off-boresight angle and the one-sided 13.8-deg Earth-limb angle at lunar distance. The colored central- and side-lobe fields are schematic and are not used as fixed angular classifiers.

c, Signal-specific transmit-pattern evidence normalized to the boresight EIRP. GPS envelopes combine public, measured NAVCEN patterns for the observed GPS IIR/IIR-M, IIF and GPS III space vehicles with equal weight per available SVN. Galileo curves use the JRC Galileo Reference Antenna Pattern (GRAP) FOC population model for E1 and E5a. For all four signals, solid colored curves show the pattern median. At each off-boresight angle, the colored P5-P95 envelope contains the central 90% of normalized-EIRP values across equally weighted satellite/azimuth pattern realizations; it is a variability envelope, not a 95% confidence interval. Green vertical bands show the full minimum-to-maximum range of retained LuGRE off-boresight-angle observations. The 13.8-deg line describes Earth-limb geometry, not a main-lobe boundary. No single fixed off-boresight threshold is imposed for either constellation.
"""
    caption_path.write_text(caption, encoding="utf-8")

    missing_l1 = ", ".join(f"SVN{value}" for value in gps_missing_svns["GPS L1"])
    missing_l5 = ", ".join(f"SVN{value}" for value in gps_missing_svns["GPS L5"])
    methods = f"""METHOD AND REFERENCE NOTES

1. LuGRE observations
Source: LuGRE first-results data products and the full 120-s geometry/noise-window table in this package. The source windows have a 60-s stride; panel a greedily retains non-overlapping windows within each operation/SV/signal sequence. Raw windows: {len(raw_windows):,}; retained windows: {len(independent_windows):,}.

2. GPS antenna patterns
Source: U.S. Coast Guard Navigation Center, GPS Technical References, including GPS IIR/IIR-M SV-specific L1/L2 patterns, GPS IIF SVN62-SVN73 L1/L2/L5 pattern workbooks, and GPS III SVN74-SVN78 L1/L2/L5 directivity files.
URL: https://navcen.uscg.gov/gps-technical-references
The 2025 PRN-to-SVN mapping is taken from the IGS satellite metadata SINEX file:
https://files.igs.org/pub/station/general/igs_satellite_metadata.snx
Patterns were available for all observed GPS IIR/IIR-M and IIF SVNs and for GPS III SVN74-SVN78. Observed GPS III patterns not public in the downloaded NAVCEN set were not replaced by another satellite: GPS L1 missing {missing_l1 or 'none'}; GPS L5 missing {missing_l5 or 'none'}. Some IIF workbooks provide conical cuts only to 23 deg; the exact angle-dependent satellite coverage is reported in lugre_main_lobe_pattern_coverage.csv.

3. Galileo antenna patterns
Source: European Commission Joint Research Centre, Galileo Reference Antenna Pattern (GRAP) model and technical report JRC135110.
Project page: https://joint-research-centre.ec.europa.eu/projects-and-activities/galileo-reference-antenna-pattern-model_en
Technical report: https://publications.jrc.ec.europa.eu/repository/bitstream/JRC135110/JRC135110_01.pdf
GRAP is a population model for the Galileo FOC constellation, not the measured pattern of one named Galileo satellite. It provides E1, E5a, E5b and E6 two-dimensional EIRP grids and model confidence bounds. It does not define 20.5 deg or 23.5 deg as a universal Galileo main-lobe edge.

4. Signal definitions
LuGRE processing uses GPS L1 C/A, GPS L5Q, Galileo E1C and Galileo E5a-Q pilot/data channels as described in the LuGRE first-results paper.
Paper: https://doi.org/10.33012/navi.756
Galileo carrier definitions follow the Galileo Open Service SIS ICD:
https://www.gsc-europa.eu/sites/default/files/sites/all/files/Galileo_OS_SIS_ICD_in_force.pdf

5. Interpretation
The Earth-limb angle (13.8 deg at lunar distance) is geometric. Main-lobe/side-lobe transitions depend on signal, space-vehicle antenna pattern and azimuth. The plotted 5-95% intervals are descriptive envelopes and never selection filters.
"""
    methods_path.write_text(methods, encoding="utf-8")
    return caption_path, methods_path


def save_figure(fig: plt.Figure) -> list[Path]:
    paths: list[Path] = []
    for stem in (OUTPUT_STEM, CURRENT_STEM):
        for suffix in (".png", ".pdf", ".svg"):
            path = stem.with_suffix(suffix)
            kwargs: dict[str, object] = {
                "bbox_inches": "tight",
                "pad_inches": 0.05,
                "facecolor": "white",
            }
            if suffix == ".png":
                kwargs["dpi"] = 600
            fig.savefig(path, **kwargs)
            paths.append(path)
    return paths


def main() -> None:
    configure_style()
    raw_windows, independent_windows = load_windows()
    prn_svn = load_prn_svn_map()

    summaries: dict[str, PatternSummary] = {}
    gps_observed_svns: dict[str, list[int]] = {}
    gps_missing_svns: dict[str, list[int]] = {}
    for signal in ("GPS L1", "GPS L5"):
        observed_prns = set(
            raw_windows.loc[raw_windows["signal_name"] == signal, "sat"].astype(str)
        )
        summary, observed_svns, missing_svns = load_gps_summary(
            signal, observed_prns, prn_svn
        )
        summaries[signal] = summary
        gps_observed_svns[signal] = observed_svns
        gps_missing_svns[signal] = missing_svns
    summaries["Galileo E1"] = load_galileo_summary("Galileo E1")
    summaries["Galileo E5a"] = load_galileo_summary("Galileo E5a")

    fig = plt.figure(figsize=(8.35, 10.15))
    outer = fig.add_gridspec(
        2,
        1,
        left=0.072,
        right=0.985,
        bottom=0.112,
        top=0.938,
        height_ratios=[1.03, 1.0],
        hspace=0.30,
    )
    _, panel_a_legend = draw_panel_a(fig, outer[0], independent_windows)
    bottom = GridSpecFromSubplotSpec(
        1,
        2,
        subplot_spec=outer[1],
        width_ratios=[0.88, 1.45],
        wspace=0.13,
    )
    draw_panel_b(fig, bottom[0])
    _, panel_c_legend = draw_panel_c(
        fig, bottom[1], summaries, independent_windows
    )

    fig.legend(
        handles=panel_a_legend,
        frameon=False,
        loc="center",
        bbox_to_anchor=(0.535, 0.548),
        ncol=6,
        columnspacing=0.9,
        handlelength=1.55,
        handletextpad=0.38,
        fontsize=5.6,
    )
    fig.legend(
        handles=panel_c_legend,
        frameon=False,
        loc="lower center",
        bbox_to_anchor=(0.705, 0.038),
        ncol=3,
        columnspacing=1.15,
        handlelength=1.8,
        handletextpad=0.42,
        labelspacing=0.45,
        fontsize=5.6,
    )

    fig.text(
        0.072,
        0.982,
        "a  LuGRE pseudorange-noise evidence by signal",
        ha="left",
        va="top",
        fontsize=10.5,
        fontweight="bold",
        color=COLORS["ink"],
    )
    fig.text(
        0.072,
        0.510,
        "b  Geometry and lobe interpretation",
        ha="left",
        va="top",
        fontsize=10.0,
        fontweight="bold",
        color=COLORS["ink"],
    )
    fig.text(
        0.435,
        0.510,
        "c  Transmit-pattern evidence by signal",
        ha="left",
        va="top",
        fontsize=10.0,
        fontweight="bold",
        color=COLORS["ink"],
    )

    paths = save_figure(fig)
    coverage_path = write_coverage(
        raw_windows,
        independent_windows,
        prn_svn,
        summaries,
        gps_observed_svns,
        gps_missing_svns,
    )
    caption_path, methods_path = write_caption_and_methods(
        raw_windows, independent_windows, gps_missing_svns
    )
    plt.close(fig)

    print("Generated:")
    for path in paths:
        print(path)
    print(coverage_path)
    print(caption_path)
    print(methods_path)
    for signal in ("GPS L1", "GPS L5"):
        summary = summaries[signal]
        print(
            f"{signal}: observed SVN={len(gps_observed_svns[signal])}; "
            f"patterned SVN={len(summary.svns)}; missing={gps_missing_svns[signal]}"
        )


if __name__ == "__main__":
    main()
